#!/usr/bin/env python3
import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Dict, List, Any

import openpyxl

try:
    from sentence_transformers import SentenceTransformer
except Exception:
    SentenceTransformer = None

try:
    import requests
except Exception:
    requests = None


DEFAULT_BASE_DIR = Path("/mnt/data")
DEFAULT_OUTPUT_JSONL = DEFAULT_BASE_DIR / "response_knowledge.jsonl"
DEFAULT_COLLECTION = "RAG_response_knowledge"
DEFAULT_QDRANT_URL = "http://localhost:6333"
DEFAULT_EMBED_MODEL = "/home/i77/AS-Platform/models/bge-small-en-v1.5"


REQUIRED_FILES = {
    "techniques": "enterprise-attack-v18.1-techniques.xlsx",
    "mitigations": "enterprise-attack-v18.1-mitigations.xlsx",
    "relationships": "enterprise-attack-v18.1-relationships.xlsx",
    "detectionstrategies": "enterprise-attack-v18.1-detectionstrategies.xlsx",
    "datacomponents": "enterprise-attack-v18.1-datacomponents.xlsx",
}


def load_sheet_rows(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path.name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    output = []

    for raw in rows[1:]:
        row = {}
        empty = True
        for idx, value in enumerate(raw):
            key = headers[idx] if idx < len(headers) else f"col_{idx}"
            if key:
                row[key] = value
            if value not in (None, ""):
                empty = False
        if not empty:
            output.append(row)

    return output


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def ensure_list(value: Any, sep: str = ",") -> List[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [clean_text(v) for v in value if clean_text(v)]
    return [clean_text(v) for v in str(value).split(sep) if clean_text(v)]


def short_hash(value: str) -> str:
    return hashlib.md5(value.encode("utf-8")).hexdigest()


def infer_response_guidance(technique: Dict[str, Any]) -> Dict[str, List[str]]:
    name = clean_text(technique.get("name"))
    description = clean_text(technique.get("description")).lower()
    platforms = ensure_list(technique.get("platforms"))
    tactics = ensure_list(technique.get("tactics"))

    recommended_actions: List[str] = []
    evidence_to_collect: List[str] = []
    containment_actions: List[str] = []

    # Universal actions
    recommended_actions.extend([
        "Validate alert context against local host telemetry and recent related alerts.",
        "Preserve volatile evidence before making destructive changes where possible.",
    ])
    evidence_to_collect.extend([
        "Parent/child process tree around the alert timestamp.",
        "Recent authentication, process, file, and network telemetry for the affected host.",
    ])

    tactic_blob = " | ".join(tactics).lower()
    platform_blob = " | ".join(platforms).lower()
    content_blob = f"{name.lower()} {description} {tactic_blob} {platform_blob}"

    if any(x in tactic_blob for x in ["credential access", "privilege escalation", "initial access"]):
        recommended_actions.extend([
            "Review account activity and reset or disable impacted credentials if compromise is suspected.",
            "Check whether the same user or UID succeeded after multiple failures.",
        ])
        evidence_to_collect.extend([
            "Authentication logs covering at least 15 minutes before and after the alert.",
            "Account privilege memberships and any recent privilege changes.",
        ])

    if "brute force" in content_blob or "password" in content_blob or "credential" in content_blob:
        containment_actions.extend([
            "Apply temporary account lockout, MFA challenge, or conditional access where appropriate.",
            "Block suspicious remote source IPs only when a real remote origin exists.",
        ])
        evidence_to_collect.append("Failed and successful logon correlation by user, UID, and source IP.")

    if "linux" in platform_blob:
        recommended_actions.extend([
            "Inspect auth.log or journalctl for correlated sudo, su, ssh, or PAM events.",
            "Review systemd services, cron jobs, shell history, and startup persistence on the endpoint.",
        ])
        evidence_to_collect.extend([
            "Output of ps auxf, ss -plant, systemctl list-units, crontab -l, and recent shell history.",
            "Hashes and timestamps for touched files under sensitive paths like /etc, /usr/local, /var/www, and /tmp.",
        ])

    if "windows" in platform_blob:
        recommended_actions.extend([
            "Inspect Security, System, PowerShell, and Sysmon logs around the alert.",
            "Review services, scheduled tasks, Run keys, and startup folders for persistence.",
        ])
        evidence_to_collect.extend([
            "Event IDs such as 4688, 4697, 7045, PowerShell operational logs, and service creation data.",
            "Running processes, autoruns, scheduled tasks, and recent registry persistence changes.",
        ])

    if "service" in content_blob:
        containment_actions.extend([
            "Stop and disable suspicious services only after capturing service configuration and binary paths.",
        ])
        evidence_to_collect.append("Service name, binary path, service start mode, and creation/change timestamps.")

    if "powershell" in content_blob or "amsi" in content_blob:
        containment_actions.extend([
            "Isolate the host if malicious PowerShell persistence or payload execution is confirmed.",
            "Quarantine or remove the malicious script or binary after evidence capture.",
        ])
        evidence_to_collect.extend([
            "Decoded command line, script block logs, AMSI-related telemetry, and dropped payload paths.",
        ])

    if "web shell" in content_blob or "/var/www/" in content_blob or ".php" in content_blob:
        containment_actions.extend([
            "Remove public access or isolate the web application host if active exploitation is confirmed.",
            "Quarantine the malicious web file after collecting a copy and hash.",
        ])
        evidence_to_collect.extend([
            "Web server access/error logs, file hashes, parent process lineage, and nearby modified web content.",
        ])

    def dedupe(items: List[str]) -> List[str]:
        seen = set()
        out = []
        for item in items:
            item = clean_text(item)
            if item and item not in seen:
                seen.add(item)
                out.append(item)
        return out

    return {
        "recommended_actions": dedupe(recommended_actions),
        "containment_actions": dedupe(containment_actions),
        "evidence_to_collect": dedupe(evidence_to_collect),
    }


def build_playbook_text(doc: Dict[str, Any]) -> str:
    parts = [
        f"Response Playbook: {doc['title']}",
        f"ATT&CK Technique: {doc.get('attack_id', '')}",
        f"Tactics: {', '.join(doc.get('tactics', []))}",
        f"Platforms: {', '.join(doc.get('platforms', []))}",
        f"Technique Description: {doc.get('description', '')}",
        f"Detection Strategies: {', '.join(doc.get('detection_strategy_ids', []))}",
        f"Mitigations: {', '.join(doc.get('mitigation_ids', []))}",
        "Recommended Actions:",
        *[f"- {x}" for x in doc.get("recommended_actions", [])],
        "Containment Actions:",
        *[f"- {x}" for x in doc.get("containment_actions", [])],
        "Evidence To Collect:",
        *[f"- {x}" for x in doc.get("evidence_to_collect", [])],
        "Procedure Examples:",
        *[f"- {x}" for x in doc.get("procedure_examples", [])[:5]],
    ]
    return "\n".join([p for p in parts if clean_text(p)])


def build_documents(base_dir: Path) -> List[Dict[str, Any]]:
    paths = {k: base_dir / v for k, v in REQUIRED_FILES.items()}
    missing = [str(p) for p in paths.values() if not p.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required files: {missing}")

    techniques = load_sheet_rows(paths["techniques"], "techniques")
    assoc_mitigations = load_sheet_rows(paths["techniques"], "associated mitigations")
    assoc_detection = load_sheet_rows(paths["techniques"], "associated detection strategies")
    procedure_examples = load_sheet_rows(paths["techniques"], "procedure examples")

    mitigations = load_sheet_rows(paths["mitigations"], "mitigations")
    techniques_addressed = load_sheet_rows(paths["mitigations"], "techniques addressed")

    relationships = load_sheet_rows(paths["relationships"], "relationships")

    detection_strategies = load_sheet_rows(paths["detectionstrategies"], "detectionstrategies")
    detection_analytic = load_sheet_rows(paths["detectionstrategies"], "detectionstrategies-analytic")

    data_components = load_sheet_rows(paths["datacomponents"], "datacomponents")

    techniques_by_id = {clean_text(r.get("ID")): r for r in techniques if clean_text(r.get("ID"))}
    mitigations_by_id = {clean_text(r.get("ID")): r for r in mitigations if clean_text(r.get("ID"))}
    detection_by_id = {clean_text(r.get("ID")): r for r in detection_strategies if clean_text(r.get("ID"))}

    technique_to_mitigations: Dict[str, List[Dict[str, Any]]] = {}
    for row in assoc_mitigations + techniques_addressed:
        tech_id = clean_text(row.get("target ID"))
        mit_id = clean_text(row.get("source ID"))
        if tech_id and mit_id:
            technique_to_mitigations.setdefault(tech_id, []).append(row)

    technique_to_detection: Dict[str, List[Dict[str, Any]]] = {}
    for row in assoc_detection:
        tech_id = clean_text(row.get("target ID"))
        det_id = clean_text(row.get("source ID"))
        if tech_id and det_id:
            technique_to_detection.setdefault(tech_id, []).append(row)

    technique_to_examples: Dict[str, List[str]] = {}
    for row in procedure_examples:
        tech_id = clean_text(row.get("target ID"))
        desc = clean_text(row.get("mapping description"))
        if tech_id and desc:
            technique_to_examples.setdefault(tech_id, []).append(desc)

    detection_to_analytics: Dict[str, List[str]] = {}
    for row in detection_analytic:
        ds_name = clean_text(row.get("detection_strategy_name"))
        analytic_name = clean_text(row.get("analytic_name"))
        if ds_name and analytic_name:
            detection_to_analytics.setdefault(ds_name, []).append(analytic_name)

    docs: List[Dict[str, Any]] = []

    for attack_id, technique in techniques_by_id.items():
        tactic_list = ensure_list(technique.get("tactics"))
        platform_list = ensure_list(technique.get("platforms"))

        mitigation_rows = technique_to_mitigations.get(attack_id, [])
        mitigation_ids = []
        mitigation_names = []
        mitigation_descriptions = []
        for row in mitigation_rows:
            mit_id = clean_text(row.get("source ID"))
            mit = mitigations_by_id.get(mit_id, {})
            if mit_id:
                mitigation_ids.append(mit_id)
            if clean_text(row.get("source name")):
                mitigation_names.append(clean_text(row.get("source name")))
            if clean_text(mit.get("description")):
                mitigation_descriptions.append(clean_text(mit.get("description")))

        detection_rows = technique_to_detection.get(attack_id, [])
        detection_ids = []
        detection_names = []
        for row in detection_rows:
            det_id = clean_text(row.get("source ID"))
            det_name = clean_text(row.get("source name"))
            if det_id:
                detection_ids.append(det_id)
            if det_name:
                detection_names.append(det_name)

        response_guidance = infer_response_guidance(technique)

        doc = {
            "id": f"response_playbook::{attack_id}",
            "doc_type": "response_playbook",
            "attack_id": attack_id,
            "title": clean_text(technique.get("name")),
            "description": clean_text(technique.get("description")),
            "attack_url": clean_text(technique.get("url")),
            "platforms": platform_list,
            "tactics": tactic_list,
            "is_sub_technique": bool(technique.get("is sub-technique")),
            "parent_technique": clean_text(technique.get("sub-technique of")),
            "mitigation_ids": sorted(set(mitigation_ids)),
            "mitigation_names": sorted(set(mitigation_names)),
            "mitigation_descriptions": list(dict.fromkeys(mitigation_descriptions))[:10],
            "detection_strategy_ids": sorted(set(detection_ids)),
            "detection_strategy_names": sorted(set(detection_names)),
            "procedure_examples": list(dict.fromkeys(technique_to_examples.get(attack_id, [])))[:8],
            "recommended_actions": response_guidance["recommended_actions"],
            "containment_actions": response_guidance["containment_actions"],
            "evidence_to_collect": response_guidance["evidence_to_collect"],
            "auto_executable": False,
            "requires_approval": True,
            "destructive": False,
            "target_platforms": platform_list,
            "source_trace": {
                "technique_sheet": "techniques",
                "associated_mitigations_sheet": "associated mitigations",
                "associated_detection_sheet": "associated detection strategies",
                "procedure_examples_sheet": "procedure examples",
            },
        }
        doc["text"] = build_playbook_text(doc)
        docs.append(doc)

    for mit_id, mitigation in mitigations_by_id.items():
        doc = {
            "id": f"mitigation::{mit_id}",
            "doc_type": "mitigation_mapping",
            "mitigation_id": mit_id,
            "title": clean_text(mitigation.get("name")),
            "description": clean_text(mitigation.get("description")),
            "url": clean_text(mitigation.get("url")),
            "text": "\n".join([
                f"Mitigation: {mit_id} - {clean_text(mitigation.get('name'))}",
                clean_text(mitigation.get("description")),
                f"Domains: {clean_text(mitigation.get('domain'))}",
            ]),
        }
        docs.append(doc)

    for det_id, det in detection_by_id.items():
        name = clean_text(det.get("name"))
        analytics = detection_to_analytics.get(name, [])
        doc = {
            "id": f"detection_strategy::{det_id}",
            "doc_type": "detection_response_map",
            "detection_strategy_id": det_id,
            "title": name,
            "description": f"Detection strategy {det_id} used to map detections to response guidance.",
            "analytics": analytics,
            "url": clean_text(det.get("url")),
            "text": "\n".join([
                f"Detection Strategy: {det_id} - {name}",
                f"Analytics: {', '.join(analytics)}",
                f"Domain: {clean_text(det.get('domain'))}",
            ]),
        }
        docs.append(doc)

    for dc in data_components:
        dc_id = clean_text(dc.get("ID"))
        if not dc_id:
            continue
        doc = {
            "id": f"data_component::{dc_id}",
            "doc_type": "evidence_collection_guide",
            "data_component_id": dc_id,
            "title": clean_text(dc.get("name")),
            "description": clean_text(dc.get("description")),
            "url": clean_text(dc.get("url")),
            "text": "\n".join([
                f"Data Component: {dc_id} - {clean_text(dc.get('name'))}",
                clean_text(dc.get("description")),
            ]),
        }
        docs.append(doc)

    rel_useful = 0
    for row in relationships:
        source_type = clean_text(row.get("source type"))
        target_type = clean_text(row.get("target type"))
        mapping_type = clean_text(row.get("mapping type"))
        if mapping_type in {"uses", "attributed-to", "mitigates", "detects"}:
            rel_useful += 1

    print(f"[INFO] Built {len(docs)} response knowledge documents.")
    print(f"[INFO] Techniques: {len(techniques_by_id)} | Mitigations: {len(mitigations_by_id)} | Detection strategies: {len(detection_by_id)} | Data components: {len(data_components)}")
    print(f"[INFO] Useful relationship rows observed for future enrichment: {rel_useful}")
    return docs


def export_jsonl(docs: List[Dict[str, Any]], output_path: Path) -> None:
    with open(output_path, "w", encoding="utf-8") as f:
        for doc in docs:
            f.write(json.dumps(doc, ensure_ascii=False) + "\n")
    print(f"[OK] Wrote JSONL to {output_path}")


def ensure_qdrant_collection(qdrant_url: str, collection: str, vector_size: int) -> None:
    if requests is None:
        raise RuntimeError("requests is required for Qdrant upload")

    url = f"{qdrant_url}/collections/{collection}"
    payload = {
        "vectors": {
            "size": vector_size,
            "distance": "Cosine",
        }
    }
    response = requests.put(url, json=payload, timeout=15)
    if response.status_code not in (200, 201):
        raise RuntimeError(f"Failed to create/update Qdrant collection: {response.status_code} {response.text}")
    print(f"[OK] Qdrant collection ready: {collection}")


def upload_to_qdrant(docs: List[Dict[str, Any]], qdrant_url: str, collection: str, embed_model: str) -> None:
    if SentenceTransformer is None:
        raise RuntimeError("sentence-transformers is required for Qdrant upload")
    if requests is None:
        raise RuntimeError("requests is required for Qdrant upload")

    model = SentenceTransformer(embed_model)
    sample_vec = model.encode("test").tolist()
    ensure_qdrant_collection(qdrant_url, collection, len(sample_vec))

    points = []
    for idx, doc in enumerate(docs):
        vector = model.encode(doc["text"]).tolist()
        payload = dict(doc)
        point_id = int(short_hash(doc["id"])[:12], 16)
        points.append({
            "id": point_id,
            "vector": vector,
            "payload": payload,
        })

        if len(points) >= 64 or idx == len(docs) - 1:
            upsert_url = f"{qdrant_url}/collections/{collection}/points?wait=true"
            response = requests.put(upsert_url, json={"points": points}, timeout=120)
            if response.status_code not in (200, 201):
                raise RuntimeError(f"Qdrant upsert failed: {response.status_code} {response.text}")
            print(f"[OK] Uploaded batch of {len(points)} docs to {collection}")
            points = []


def parse_args():
    parser = argparse.ArgumentParser(description="Build ATT&CK-based response knowledge documents.")
    parser.add_argument("--base-dir", default=str(DEFAULT_BASE_DIR), help="Directory containing the ATT&CK xlsx files.")
    parser.add_argument("--output-jsonl", default=str(DEFAULT_OUTPUT_JSONL), help="Output JSONL path.")
    parser.add_argument("--upload-qdrant", action="store_true", help="Also embed and upload docs to Qdrant.")
    parser.add_argument("--qdrant-url", default=DEFAULT_QDRANT_URL, help="Qdrant base URL.")
    parser.add_argument("--collection", default=DEFAULT_COLLECTION, help="Qdrant collection name.")
    parser.add_argument("--embed-model", default=DEFAULT_EMBED_MODEL, help="SentenceTransformer model path.")
    return parser.parse_args()


def main():
    args = parse_args()
    base_dir = Path(args.base_dir)
    output_jsonl = Path(args.output_jsonl)

    docs = build_documents(base_dir=base_dir)
    export_jsonl(docs, output_jsonl)

    if args.upload_qdrant:
        upload_to_qdrant(
            docs=docs,
            qdrant_url=args.qdrant_url,
            collection=args.collection,
            embed_model=args.embed_model,
        )


if __name__ == "__main__":
    main()
